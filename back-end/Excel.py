from openpyxl import load_workbook
import datetime

def get_data(p_file: str, p_worksheet =None, p_columns: list =None, p_header: bool =False, p_first_row: int =1):
    """
    Read the data from excel

    :param p_file: file path and file name
    :param p_worksheet: worksheet name. If worksheet isn't defined, all worksheets will be read
    :param p_columns: list of columns. If columns aren't defined, only the first cell will be read
    :param p_first_row: the number of a table first row
    :param p_header: table has a header or not
    :return: tuple of data or error
    """
    try:
        l_datetime=str(datetime.datetime.now())
        l_workbook=load_workbook(filename = p_file, read_only=True)
        if not p_worksheet:
            l_worksheet_name_list=l_workbook.sheetnames
        else:
            l_worksheet_name_list=[p_worksheet]
        # transform the data into list of tuples
        l_data=[]
        for i_worksheet_name in l_worksheet_name_list: # read the worksheet/all worksheets
            l_worksheet=l_workbook[i_worksheet_name]
            if not p_columns: # if column aren't defined, the first cell will be read. Can be used to check a file/worksheet existing
                return l_worksheet['A1'].value, None
            # determine the table in a worksheet
            l_table=l_worksheet[p_first_row:l_worksheet.max_row]
            if type(l_table[0])!=tuple:
                l_table_data=[l_table]
            else:
                l_table_data=l_table
            # determine the value of every cell in the first row
            l_cell_number_dict={}
            for i, i_cell in enumerate(l_table_data[0]):
                l_cell_number_dict.update({i_cell.value:i}) # dictionary with value of a cell and cell's number
            for i, i_row in enumerate(l_table_data):
                if p_header and i==0: # if there is a header, skip the row
                    continue
                l_row=[]
                for i_column in p_columns:
                    if i_column.isdigit(): # if column name is a number, it's the number of a column in a table
                        if int(i_column)>i_row.__len__(): # if the number of a column is bigger than the number of the last column
                            return None, 'There is no '+str(i_column)+'-th column'
                        l_row.append(i_row[int(i_column)-1].value)
                    else: # if column name is the name of a column in a header
                        l_cell_number=l_cell_number_dict.get(i_column)
                        if l_cell_number is None: # if there is no such column in a header
                            return None, 'There is no '+i_column+" column"
                        else:
                            l_row.append(i_row[l_cell_number].value)
                # add current_timestamp value
                l_row.append(l_datetime)
                l_data.append(tuple(l_row)) # transofrm the list into list of tuples
        #close the workbook
        l_workbook.close()
        return l_data, None
    except Exception as e:
        l_error=str(e)
        return None, l_error
